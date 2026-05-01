# VirVentures ASIN Verifier — v7.0 Production
# Features: Gemini AI · UPC Search · Pack Size Detection · Unit-Aware Matching
# Semantic ML · FBA/FBM Detection · Price Ranges · SQLite Cache · Concurrency

import streamlit as st
import pandas as pd
import requests
import re
import random
import time
import io
import base64
import os
import json
import sqlite3
import logging
import traceback
import threading
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from difflib import SequenceMatcher
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Gemini API
try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

# Sentence-transformers
try:
    from sentence_transformers import SentenceTransformer
    import numpy as np
    SEMANTIC_AVAILABLE = True
except ImportError:
    SEMANTIC_AVAILABLE = False

st.set_page_config(
    page_title="VirVentures ASIN Verifier v7.0",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =============================================================================
# CONSTANTS
# =============================================================================
CACHE_DB = "asin_cache.db"
ERROR_LOG = "vv_errors.log"
CACHE_TTL_HOURS = 4
DEFAULT_WORKERS = 3
MAX_WORKERS = 5
RETRY_DELAYS = [5, 10, 20]
SEMANTIC_MODEL_ID = "all-MiniLM-L6-v2"
KW_WEIGHT = 0.40
SEM_WEIGHT = 0.60

DELAY_SEQ = [8.0, 8.5, 9.8, 7.9, 10.2, 8.7, 11.1, 9.4, 7.6, 10.8, 8.2, 9.1]

UA_POOL = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36 Edg/121.0.0.0",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:123.0) Gecko/20100101 Firefox/123.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:124.0) Gecko/20100101 Firefox/124.0",
]

LANGS = ["en-US,en;q=0.9", "en-GB,en;q=0.9", "en-CA,en;q=0.8,fr;q=0.7", "en-AU,en;q=0.9"]

STOPWORDS = {
    "a","an","the","and","or","for","of","in","to","with","by","is","it","its",
    "–","-","&","at","on","from","as","are","be","this","that","will","has",
    "have","not","but","can","pk","pcs","set","new","use","used","each","per",
    "pack","count","size","color","colour","oz","lb","lbs","fl","ct","qty",
}

# Unit categories for measurement matching
UNIT_CATEGORIES = {
    'volume': ['ml', 'milliliter', 'millilitre', 'mls', 'cc', 'oz', 'fl oz', 'liter', 'ltr', 'gallon'],
    'weight': ['g', 'gram', 'grams', 'kg', 'kilogram', 'lb', 'pound', 'oz wt'],
    'count': ['each', 'piece', 'pc', 'pcs', 'count', 'unit', 'pack', 'roll', 'sheet',
              'pen', 'pens', 'marker', 'markers', 'stick', 'sticks', 'cartridge'],
    'length': ['mm', 'cm', 'm', 'inch', 'inches', 'ft', 'feet'],
}

MEASUREMENT_PATTERNS = [
    (r'(\d+(?:\.\d+)?)\s*(ml|milliliter|millilitre|mls)', 'volume'),
    (r'(\d+(?:\.\d+)?)\s*(g|gram|grams|grammes)', 'weight'),
    (r'(\d+(?:\.\d+)?)\s*(?:-?\s*)(pack|count|pieces?|pcs?|units?|each|roll|caps?|cartridge)', 'count'),
    (r'(\d+(?:\.\d+)?)\s*(inch|inches|cm|mm|m|meter)', 'length'),
    (r'(\d+(?:\.\d+)?)\s*(oz|ounce|ounces)', 'weight'),
    (r'(\d+(?:\.\d+)?)\s*(pen|pens|marker|markers|stick|sticks)', 'count'),
]

# =============================================================================
# LOGGING
# =============================================================================
logging.basicConfig(
    filename=ERROR_LOG,
    level=logging.ERROR,
    format="%(asctime)s | %(levelname)s | %(message)s",
)

def log_error(asin: str, msg: str, exc: Exception = None):
    detail = f"ASIN={asin} | {msg}"
    if exc:
        detail += f" | {traceback.format_exc()}"
    logging.error(detail)

# =============================================================================
# LOGO & CSS
# =============================================================================
def get_logo_b64():
    for p in ["virventures_logo.jpg", "virventures_com_logo.jpg"]:
        if os.path.exists(p):
            with open(p, "rb") as f:
                return base64.b64encode(f.read()).decode()
    return None

LOGO_B64 = get_logo_b64()
LOGO_HTML = (
    f'<img src="data:image/jpeg;base64,{LOGO_B64}" '
    f'style="height:64px;width:auto;border-radius:10px;flex-shrink:0;'
    f'background:#fff;padding:6px;box-shadow:0 2px 12px rgba(0,0,0,0.2);">'
    if LOGO_B64 else ""
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');
*, html, body, [class*="css"] { font-family: 'Inter', sans-serif !important; }
.stApp, .main, .block-container { background: #f8f9fb !important; }
.vv-header {
    background: linear-gradient(135deg, #1e2d4e 0%, #2a3f6e 100%);
    border-radius: 16px; padding: 26px 36px; margin-bottom: 24px;
    display: flex; align-items: center; gap: 20px;
    box-shadow: 0 8px 32px rgba(30,45,78,0.15); border-left: 6px solid #f47920;
}
.vv-title { color: #ffffff !important; font-size: 1.2rem !important; font-weight: 800 !important; margin: 0 0 8px 0 !important; }
.vv-sub { color: #f47920 !important; font-size: 0.76rem !important; font-weight: 700 !important; letter-spacing: 1.6px !important; text-transform: uppercase !important; background: rgba(244,121,32,0.15); display: inline-block; padding: 4px 14px; border-radius: 20px; }
section[data-testid="stSidebar"] { background: #ffffff !important; border-right: 2px solid #f0f0f0 !important; }
.stButton > button { background: linear-gradient(90deg, #f47920, #ff9a45) !important; color: #fff !important; font-weight: 800 !important; border: none !important; border-radius: 10px !important; padding: 14px 36px !important; }
.stDownloadButton > button { background: #fff !important; color: #f47920 !important; border: 2px solid #f47920 !important; font-weight: 700 !important; border-radius: 10px !important; }
div[data-testid="metric-container"] { background: #fff !important; border: 1.5px solid #f0f0f0 !important; border-radius: 12px !important; padding: 18px !important; }
.stProgress > div > div { background: linear-gradient(90deg, #f47920, #ffb347) !important; border-radius: 6px !important; }
.sec { color: #1e2d4e; font-size: 1rem; font-weight: 800; padding-bottom: 6px; border-bottom: 3px solid #f47920; display: inline-block; margin: 20px 0 14px 0; }
.info-box { background: #fff8f3; border-left: 4px solid #f47920; border-radius: 0 10px 10px 0; padding: 12px 16px; margin: 10px 0; }
.good-box { background: #f0faf4; border-left: 4px solid #2e7d32; border-radius: 0 10px 10px 0; padding: 12px 16px; margin: 10px 0; }
.warn-box { background: #fffbea; border-left: 4px solid #f5a623; border-radius: 0 10px 10px 0; padding: 12px 16px; margin: 10px 0; }
.worker-card { background: #fff; border: 1.5px solid #f0f0f0; border-radius: 10px; padding: 12px 16px; margin: 4px 0; font-size: 12px; }
</style>
""", unsafe_allow_html=True)

st.markdown(f"""
<div class="vv-header">
    {LOGO_HTML}
    <div>
        <p class="vv-title">Hi VirVentures 👋 — ASIN Verification v7.0</p>
        <p class="vv-sub">🔍 Gemini AI · UPC Search · Pack Size Detection · Unit-Aware Matching · P&amp;L</p>
    </div>
</div>
""", unsafe_allow_html=True)

# =============================================================================
# DATABASE / CACHE
# =============================================================================
def db_init():
    conn = sqlite3.connect(CACHE_DB)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS asin_cache (
            asin TEXT PRIMARY KEY, price_data TEXT, desc_data TEXT,
            fulfillment TEXT, fetched_at TEXT, status TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS failed_asins (
            asin TEXT PRIMARY KEY, reason TEXT, attempts INTEGER DEFAULT 1, last_tried TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS upc_cache (
            upc TEXT PRIMARY KEY, asin TEXT, pack_size INTEGER, title TEXT, fetched_at TEXT
        )
    """)
    conn.commit()
    conn.close()

def cache_get(asin: str, ttl_hours: int = CACHE_TTL_HOURS):
    try:
        conn = sqlite3.connect(CACHE_DB)
        c = conn.cursor()
        c.execute("SELECT price_data, desc_data, fulfillment, fetched_at FROM asin_cache WHERE asin=?", (asin,))
        row = c.fetchone()
        conn.close()
        if not row:
            return None
        fetched_at = datetime.fromisoformat(row[3])
        if datetime.now() - fetched_at > timedelta(hours=ttl_hours):
            return None
        return {"price_data": json.loads(row[0]) if row[0] else {},
                "desc_data": json.loads(row[1]) if row[1] else {},
                "fulfillment": json.loads(row[2]) if row[2] else {},
                "fetched_at": row[3], "from_cache": True}
    except Exception:
        return None

def cache_set(asin: str, price_data: dict, desc_data: dict, fulfillment: dict):
    try:
        conn = sqlite3.connect(CACHE_DB)
        c = conn.cursor()
        c.execute("INSERT OR REPLACE INTO asin_cache VALUES (?, ?, ?, ?, ?, 'ok')",
                  (asin, json.dumps(price_data), json.dumps(desc_data),
                   json.dumps(fulfillment), datetime.now().isoformat()))
        conn.commit()
        conn.close()
    except Exception as e:
        log_error(asin, "cache_set failed", e)

def cache_mark_failed(asin: str, reason: str):
    try:
        conn = sqlite3.connect(CACHE_DB)
        c = conn.cursor()
        c.execute("INSERT OR REPLACE INTO failed_asins VALUES (?, ?, 1, ?)",
                  (asin, reason, datetime.now().isoformat()))
        conn.commit()
        conn.close()
    except Exception:
        pass

def cache_get_failed_asins():
    try:
        conn = sqlite3.connect(CACHE_DB)
        c = conn.cursor()
        c.execute("SELECT asin FROM failed_asins ORDER BY last_tried DESC")
        rows = c.fetchall()
        conn.close()
        return [r[0] for r in rows]
    except Exception:
        return []

def cache_clear_failed(asins: list):
    try:
        conn = sqlite3.connect(CACHE_DB)
        c = conn.cursor()
        c.executemany("DELETE FROM failed_asins WHERE asin=?", [(a,) for a in asins])
        conn.commit()
        conn.close()
    except Exception:
        pass

def cache_upc_get(upc: str):
    try:
        conn = sqlite3.connect(CACHE_DB)
        c = conn.cursor()
        c.execute("SELECT asin, pack_size, title, fetched_at FROM upc_cache WHERE upc=?", (upc,))
        row = c.fetchone()
        conn.close()
        if row and datetime.now() - datetime.fromisoformat(row[3]) < timedelta(hours=168):  # 7 days
            return {"asin": row[0], "pack_size": row[1], "title": row[2]}
        return None
    except Exception:
        return None

def cache_upc_set(upc: str, asin: str, pack_size: int, title: str):
    try:
        conn = sqlite3.connect(CACHE_DB)
        c = conn.cursor()
        c.execute("INSERT OR REPLACE INTO upc_cache VALUES (?, ?, ?, ?, ?)",
                  (upc, asin, pack_size, title, datetime.now().isoformat()))
        conn.commit()
        conn.close()
    except Exception:
        pass

db_init()

# =============================================================================
# FEATURE 1: GEMINI API INTEGRATION (Free AI)
# =============================================================================
@st.cache_data(ttl=3600, show_spinner=False)
def ai_analyze_mismatch(vendor_desc: str, amazon_desc: str, combined_score: float) -> dict:
    """Use Google Gemini to analyze description mismatches (cached, free tier)"""
    if not st.session_state.get("gemini_configured", False) or not GEMINI_AVAILABLE:
        return {"insight": "Add Gemini API key in sidebar", "verdict": "N/A", "action": "REVIEW"}
    
    try:
        model = genai.GenerativeModel('gemini-2.0-flash-lite')
        prompt = f"""
        ASIN Verification AI Analysis for FBA Reseller.

        VENDOR: {vendor_desc[:600]}
        AMAZON: {amazon_desc[:600]}
        SIMILARITY SCORE: {combined_score:.0%}

        Answer EXACTLY in this format (3 lines, pipe-separated):
        VERDICT|ACTION|INSIGHT

        VERDICT: YES (same product) / MISMATCH (different) / UNCLEAR
        ACTION: LIST / REVIEW / SKIP
        INSIGHT: Very brief reason (max 80 chars)

        Example:
        YES|LIST|Product matches, good margin
        MISMATCH|SKIP|Different quantities (18ml vs 24-pack)
        """
        response = model.generate_content(prompt)
        parts = response.text.strip().split('|')
        
        return {
            "verdict": parts[0] if len(parts) > 0 else "UNCLEAR",
            "action": parts[1] if len(parts) > 1 else "REVIEW",
            "insight": parts[2] if len(parts) > 2 else "AI analysis complete"
        }
    except Exception as e:
        return {"insight": f"API Error", "verdict": "ERROR", "action": "REVIEW"}

# =============================================================================
# FEATURE 2: UPC SEARCH WITH PACK SIZE DETECTION
# =============================================================================
def detect_pack_size(title: str) -> int:
    """Extract pack/count from product title. Returns 1 for single items."""
    title_lower = title.lower()
    
    patterns = [
        r'(\d+)\s*-pack', r'(\d+)\s*pack', r'(\d+)\s*count', r'(\d+)\s*pieces?',
        r'(\d+)\s*units?', r'(\d+)\s*ct', r'pack of (\d+)', r'(\d+)\s*[\|/]\s*pack',
        r'(\d+)\s*-?pcs?', r'(\d+)\s*-?set'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, title_lower)
        if match:
            count = int(match.group(1))
            if 1 <= count <= 100:
                return count
    
    singular_indicators = ['each', 'single', '1-', '-1 ', '1 piece', '1 unit', 'individual']
    for ind in singular_indicators:
        if ind in title_lower:
            return 1
    
    return 1  # Default to single

def search_asins_by_upc(upc: str, session: requests.Session) -> list:
    """Search Amazon for all ASINs associated with a UPC. Returns list of candidates."""
    if not upc or len(upc) < 8:
        return []
    
    clean_upc = re.sub(r'[^\d]', '', upc)
    results = []
    
    # Check cache first
    cached = cache_upc_get(clean_upc)
    if cached:
        return [cached]
    
    search_url = f"https://www.amazon.com/s?k={clean_upc}&i=merchant-items"
    
    try:
        r = session.get(search_url, headers={
            "User-Agent": random.choice(UA_POOL),
            "Accept-Language": random.choice(LANGS),
        }, timeout=15)
        
        if r.status_code != 200:
            return []
        
        soup = BeautifulSoup(r.text, "lxml")
        
        for result in soup.select("[data-asin], .s-result-item"):
            asin = result.get("data-asin")
            if not asin or len(asin) != 10:
                continue
            
            title_tag = result.select_one("h2 a span, .a-size-base-plus")
            title = title_tag.get_text(strip=True) if title_tag else ""
            
            pack_size = detect_pack_size(title)
            
            price_tag = result.select_one(".a-price-whole")
            price = extract_price_float(price_tag.get_text()) if price_tag else None
            
            results.append({
                'asin': asin,
                'title': title[:200],
                'pack_size': pack_size,
                'price': price,
                'source': 'search'
            })
            
            # Cache the result
            cache_upc_set(clean_upc, asin, pack_size, title[:200])
            
    except Exception as e:
        log_error(upc, f"UPC search failed", e)
    
    return results

def find_single_item_asn(upc: str, session: requests.Session) -> dict:
    """Find the single-item ASIN for a given UPC (pack_size=1)."""
    results = search_asins_by_upc(upc, session)
    
    if not results:
        return {"asin": None, "confidence": "NONE", "reason": "No ASINs found for this UPC", "pack_size": None}
    
    single_items = [r for r in results if r.get('pack_size', 1) == 1]
    
    if len(single_items) == 1:
        return {
            'asin': single_items[0]['asin'],
            'title': single_items[0]['title'],
            'confidence': 'HIGH',
            'reason': 'Single-item ASIN found via UPC lookup',
            'pack_size': 1
        }
    elif len(single_items) > 1:
        best = min(single_items, key=lambda x: x.get('price') or float('inf'))
        return {
            'asin': best['asin'],
            'title': best['title'],
            'confidence': 'MEDIUM',
            'reason': f'Multiple single items found, selecting lowest price option',
            'pack_size': 1
        }
    else:
        smallest = min(results, key=lambda x: x.get('pack_size', 999))
        return {
            'asin': smallest['asin'],
            'title': smallest['title'],
            'confidence': 'LOW',
            'reason': f'No single item found. Using smallest pack: {smallest["pack_size"]}-pack',
            'pack_size': smallest.get('pack_size', 0)
        }

# =============================================================================
# FEATURE 3: UNIT-AWARE MEASUREMENT MATCHING
# =============================================================================
def extract_measurements(text: str) -> list:
    """Extract all measurements (value, unit, category) from text."""
    measurements = []
    text_lower = text.lower()
    
    for pattern, category in MEASUREMENT_PATTERNS:
        matches = re.findall(pattern, text_lower)
        for match in matches:
            try:
                value = float(match[0]) if isinstance(match, tuple) else float(match)
                unit = match[1] if isinstance(match, tuple) and len(match) > 1 else ""
                measurements.append((value, unit, category))
            except (ValueError, TypeError):
                continue
    
    return measurements

def compare_measurements(vendor_desc: str, amazon_desc: str) -> dict:
    """Compare measurements between vendor and Amazon descriptions."""
    vendor_meas = extract_measurements(vendor_desc)
    amazon_meas = extract_measurements(amazon_desc)
    
    result = {
        'match': False,
        'vendor_measurements': vendor_meas,
        'amazon_measurements': amazon_meas,
        'confidence': 0.0,
        'reason': ''
    }
    
    if not vendor_meas and not amazon_meas:
        result['confidence'] = 0.5
        result['reason'] = 'No measurements found in either description'
        return result
    
    if vendor_meas and not amazon_meas:
        result['confidence'] = 0.2
        result['reason'] = f'Vendor has {len(vendor_meas)} measurements, Amazon has none'
        return result
    
    matches = 0
    total = len(vendor_meas)
    
    for v_val, v_unit, v_cat in vendor_meas:
        for a_val, a_unit, a_cat in amazon_meas:
            if v_cat == a_cat:
                tolerance = 0.10
                if abs(v_val - a_val) / max(v_val, 0.01) <= tolerance:
                    matches += 1
                    break
                else:
                    result['reason'] = f'Value mismatch: {v_val}{v_unit} vs {a_val}{a_unit}'
    
    result['confidence'] = matches / total if total > 0 else 0
    result['match'] = result['confidence'] >= 0.8
    
    if result['match']:
        result['reason'] = f'Measurements match: {matches}/{total}'
    elif result['confidence'] > 0:
        result['reason'] = f'Partial measurement match: {matches}/{total}'
    
    return result

def enhanced_description_match(vendor_desc: str, amazon_desc: str, sem_model) -> dict:
    """Combined semantic matching + measurement comparison."""
    # Original semantic match
    kw_score, ratio_str, matched, missed = keyword_match_score(vendor_desc, amazon_desc)
    
    if sem_model is not None and SEMANTIC_AVAILABLE:
        sem_score = semantic_similarity(vendor_desc[:512], amazon_desc[:512], sem_model)
        semantic_combined = (KW_WEIGHT * kw_score) + (SEM_WEIGHT * sem_score)
    else:
        sem_score = 0.0
        semantic_combined = kw_score
    
    # Measurement comparison
    measurement_result = compare_measurements(vendor_desc, amazon_desc)
    
    # Weighted final score: 70% semantic + 30% measurement
    # BUT: Critical mismatch penalizes heavily
    if measurement_result['confidence'] < 0.3 and measurement_result['vendor_measurements']:
        final_score = min(semantic_combined * 0.3, 0.2)
        alert_level = "CRITICAL"
        reason = f"⚠️ MEASUREMENT MISMATCH: {measurement_result['reason']}"
    elif measurement_result['confidence'] > 0:
        final_score = (semantic_combined * 0.7) + (measurement_result['confidence'] * 0.3)
        alert_level = "INFO"
        reason = f"✅ Measurements: {measurement_result['reason']}"
    else:
        final_score = semantic_combined
        alert_level = "INFO"
        reason = "No measurements to compare"
    
    return {
        'final_score': final_score,
        'semantic_score': semantic_combined,
        'measurement_score': measurement_result['confidence'],
        'measurement_match': measurement_result['match'],
        'vendor_measurements': str(measurement_result['vendor_measurements'])[:100],
        'amazon_measurements': str(measurement_result['amazon_measurements'])[:100],
        'reason': reason,
        'alert_level': alert_level,
        'recommendation': '✅ MATCH' if final_score > 0.7 else '⚠️ VERIFY' if final_score > 0.4 else '❌ MISMATCH'
    }

# =============================================================================
# EXISTING FUNCTIONS (scraping, price extraction, etc.)
# =============================================================================
def build_session(cookies: dict = None):
    session = requests.Session()
    if cookies:
        for name, value in cookies.items():
            session.cookies.set(name, value, domain=".amazon.com")
    return session

def random_headers():
    return {"User-Agent": random.choice(UA_POOL), "Accept-Language": random.choice(LANGS),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Encoding": "gzip, deflate, br", "Connection": "keep-alive", "DNT": "1"}

def is_captcha(text: str):
    return any(x in text for x in ["api-services-support@amazon.com", "Enter the characters you see below",
                                    "automated access", "Type the characters you see in this image"])

def fetch_asin_page(asin: str, session: requests.Session, max_retries: int = 3):
    urls = [f"https://www.amazon.com/dp/{asin}", f"https://www.amazon.com/gp/product/{asin}"]
    for attempt in range(max_retries):
        url = urls[attempt % len(urls)]
        delay = RETRY_DELAYS[min(attempt, len(RETRY_DELAYS)-1)]
        try:
            r = session.get(url, headers=random_headers(), timeout=22)
            if r.status_code == 200 and not is_captcha(r.text):
                return BeautifulSoup(r.text, "lxml"), "OK"
            elif r.status_code in (429, 503) and attempt < max_retries - 1:
                time.sleep(delay + random.uniform(5, 12))
        except Exception as e:
            log_error(asin, f"fetch attempt {attempt+1} failed", e)
            if attempt < max_retries - 1:
                time.sleep(delay)
    return None, "FAILED"

def fetch_offer_listing(asin: str, session: requests.Session):
    url = f"https://www.amazon.com/gp/offer-listing/{asin}/ref=dp_olp_all_mbc?ie=UTF8&condition=new"
    try:
        r = session.get(url, headers=random_headers(), timeout=18)
        if r.status_code == 200 and not is_captcha(r.text):
            return BeautifulSoup(r.text, "lxml")
    except Exception as e:
        log_error(asin, "offer listing fetch failed", e)
    return None

def extract_price_float(text: str):
    cleaned = re.sub(r"[^\d.]", "", text.replace(",", ""))
    try:
        val = float(cleaned)
        return val if 0.50 < val < 50000 else None
    except (ValueError, TypeError):
        return None

def detect_fulfillment_type(container_html: str):
    text_lower = container_html.lower()
    if "fulfilled by amazon" in text_lower or "amazon fulfillment" in text_lower:
        return "FBA"
    if "fulfilled by merchant" in text_lower:
        return "FBM"
    return "UNKNOWN"

def extract_all_prices(soup_product, soup_offers, your_bb, seller_type="FBA"):
    result = {"min_price": None, "max_price": None, "selected_price": None,
              "price_range_str": "N/A", "selection_reason": "", "fba_prices": [],
              "fbm_prices": [], "all_offers": [], "bb_price": None}
    all_prices_raw = []
    
    BB_SELECTORS = ["#corePriceDisplay_desktop_feature_div .a-price-whole",
                    "#corePriceDisplay_desktop_feature_div .a-offscreen",
                    "#price_inside_buybox", "#priceblock_ourprice", "#priceblock_dealprice"]
    for sel in BB_SELECTORS:
        tag = soup_product.select_one(sel)
        if tag:
            p = extract_price_float(tag.get_text(strip=True))
            if p:
                result["bb_price"] = p
                all_prices_raw.append({"price": p, "fulfillment": "FBA"})
                break
    
    if soup_offers:
        for offer_row in soup_offers.select(".olpOffer, .a-row.olpOffer")[:20]:
            row_html = str(offer_row)
            fulfill = detect_fulfillment_type(row_html)
            price_tag = offer_row.select_one(".olpOfferPrice, .a-price .a-offscreen")
            if price_tag:
                p = extract_price_float(price_tag.get_text(strip=True))
                if p:
                    all_prices_raw.append({"price": p, "fulfillment": fulfill})
                    if fulfill == "FBA":
                        result["fba_prices"].append(p)
                    elif fulfill == "FBM":
                        result["fbm_prices"].append(p)
    
    all_valid = [e["price"] for e in all_prices_raw if e.get("price") and 0.50 < e["price"] < 50000]
    if not all_valid:
        return result
    
    all_valid = sorted(set(all_valid))
    result["min_price"] = all_valid[0]
    result["max_price"] = all_valid[-1]
    
    if len(all_valid) > 1:
        result["price_range_str"] = f"${all_valid[0]:.2f} – ${all_valid[-1]:.2f}"
    else:
        result["price_range_str"] = f"${all_valid[0]:.2f}"
    
    if seller_type == "FBA" and result["fba_prices"]:
        result["selected_price"] = min(result["fba_prices"])
        result["selection_reason"] = f"Lowest FBA price (ignoring FBM)"
    elif result["bb_price"]:
        result["selected_price"] = result["bb_price"]
        result["selection_reason"] = "Buy Box price"
    else:
        result["selected_price"] = result["min_price"]
        result["selection_reason"] = "Minimum market price"
    
    return result

def extract_product_text(soup):
    data = {"title": "", "bullets": [], "description": "", "brand": "", "full_text": ""}
    title_tag = soup.select_one("#productTitle")
    if title_tag:
        data["title"] = title_tag.get_text(strip=True)
    for b in soup.select("#feature-bullets li span.a-list-item"):
        text = b.get_text(strip=True)
        if text and len(text) > 5:
            data["bullets"].append(text)
    desc_tag = soup.select_one("#productDescription")
    if desc_tag:
        data["description"] = desc_tag.get_text(" ", strip=True)
    data["full_text"] = " ".join(filter(None, [data["title"], " ".join(data["bullets"]), data["description"]])).lower()
    return data

@st.cache_resource(show_spinner="🧠 Loading ML model...")
def load_semantic_model():
    if not SEMANTIC_AVAILABLE:
        return None
    try:
        return SentenceTransformer(SEMANTIC_MODEL_ID)
    except Exception as e:
        log_error("model", "Failed to load model", e)
        return None

def semantic_similarity(text_a: str, text_b: str, model):
    if model is None or not text_a or not text_b:
        return 0.0
    try:
        emb_a = model.encode(text_a[:512], convert_to_numpy=True)
        emb_b = model.encode(text_b[:512], convert_to_numpy=True)
        return float(np.dot(emb_a, emb_b) / (np.linalg.norm(emb_a) * np.linalg.norm(emb_b) + 1e-8))
    except Exception:
        return 0.0

def normalize_word(w: str):
    w = w.lower().strip()
    if len(w) > 4:
        if w.endswith("ies"): return w[:-3] + "y"
        if w.endswith("s"): return w[:-1]
    return w

def fuzzy_word_in_text(word: str, text: str, threshold: float = 0.82):
    norm = normalize_word(word)
    if word in text or norm in text:
        return True
    if len(word) > 4:
        for tw in text.split():
            if len(tw) > 3 and SequenceMatcher(None, norm, tw).ratio() >= threshold:
                return True
    return False

def keyword_match_score(our_text: str, amz_text: str):
    if not our_text or not amz_text:
        return 0.0, "0/0", [], []
    words = list({w.lower() for w in re.findall(r"[a-zA-Z0-9]+", str(our_text))
                  if w.lower() not in STOPWORDS and len(w) > 2})
    if not words:
        return 0.0, "0/0", [], []
    matched = [w for w in words if fuzzy_word_in_text(w, amz_text)]
    missed = [w for w in words if not fuzzy_word_in_text(w, amz_text)]
    return len(matched) / len(words), f"{len(matched)}/{len(words)}", matched, missed

def recalculate_pl(selected_price, max_price, net_price, fulfillment_cost, referral_rate=0.15):
    result = {"new_breakeven": None, "conservative_profit": None, "upside_profit": None,
              "conservative_margin": None, "upside_margin": None, "remark": "", "status": "UNKNOWN", "recommendation": ""}
    if selected_price is None or net_price is None or net_price <= 0:
        result["remark"] = "Insufficient data"
        return result
    ref_fee = round(selected_price * referral_rate, 2)
    fixed_costs = fulfillment_cost or 0.0
    new_breakeven = round(net_price + fixed_costs + ref_fee, 2)
    result["new_breakeven"] = new_breakeven
    cons_profit = round(selected_price - new_breakeven, 2)
    cons_margin = round((cons_profit / selected_price) * 100, 2) if selected_price > 0 else 0
    result["conservative_profit"] = cons_profit
    result["conservative_margin"] = cons_margin
    result["status"] = "PROFITABLE" if cons_profit > 1.0 else "MARGINAL" if cons_profit > 0 else "LOSS"
    result["recommendation"] = "✅ Good to list" if cons_profit > 1 else "⚠️ Review" if cons_profit > 0 else "❌ Do not list"
    result["remark"] = f"Selected: ${selected_price:.2f} | Breakeven: ${new_breakeven:.2f} | Profit: ${cons_profit:.2f} ({cons_margin:.1f}%)"
    return result

def compute_confidence(combined_score, bb_severity, price_available, has_range, fetch_ok):
    result = {"level": "LOW", "action": "RE-EVALUATE", "reasons": []}
    if not fetch_ok:
        return result
    score = 0
    if combined_score >= 0.80: score += 50
    elif combined_score >= 0.55: score += 30
    else: score += 10
    if price_available and bb_severity == "ok": score += 30
    elif price_available: score += 15
    if has_range: score -= 5
    result["level"] = "HIGH" if score >= 70 else "MEDIUM" if score >= 40 else "LOW"
    result["action"] = "NONE" if score >= 70 else "REVIEW" if score >= 40 else "RE-EVALUATE"
    return result

# =============================================================================
# COLUMN DETECTION
# =============================================================================
ALIASES = {
    "asin": ["output asin", "asin", "input_asin", "amazon asin"],
    "title": ["title", "product name", "item name"],
    "desc": ["description", "product description", "item description"],
    "brand": ["brand", "brand name", "manufacturer"],
    "upc": ["upc", "upc#", "input_upc", "ean", "barcode"],
    "bb_price": ["bb price", "buy box price", "buybox price", "bb_price"],
    "net_price": ["net price", "netprice", "cost", "vendor cost"],
    "fulfillment": ["fulfillment cost", "fba fees", "fullfilment cost"],
    "referral": ["amazon referral fee", "referral fee"],
}

def detect_col(key: str, cols: list):
    cl = {c.strip().lower(): c for c in cols}
    for alias in ALIASES[key]:
        for k, v in cl.items():
            if alias in k or k in alias:
                return v
    return None

def parse_price(val):
    if not val or str(val).strip().lower() in ("nan", "", "none", "n/a"):
        return None
    try:
        return float(re.sub(r"[^\d.]", "", str(val)))
    except (ValueError, TypeError):
        return None

# =============================================================================
# MAIN VERIFICATION FUNCTION (UPDATED with UPC + Enhanced Matching)
# =============================================================================
def verify_single_asin(row_data, session, sem_model, config, force_refresh=False):
    upc = row_data.get("upc", "").strip()
    asin_from_sheet = row_data.get("asin", "").strip()
    our_title = row_data.get("title", "")
    our_desc = row_data.get("desc", "")
    our_brand = row_data.get("brand", "")
    your_bb = row_data.get("your_bb")
    net_price = row_data.get("net_price")
    full_cost = row_data.get("full_cost")
    
    our_text = " ".join(filter(None, [our_title, our_desc, our_brand]))
    
    base = {
        "ASIN Used": asin_from_sheet, "ASIN Source": "Sheet", "Pack Note": "",
        "UPC Provided": upc if upc else "N/A", "Price Range": "N/A", "Selected Price": "N/A",
        "FBA Prices": "", "FBM Prices": "", "BB Price (Live)": "N/A", "BB vs Sheet": "—",
        "New Breakeven": "—", "Conservative Profit": "—", "Upside Profit": "—",
        "Conservative Margin": "—", "P&L Status": "UNKNOWN", "Recommendation": "",
        "Amazon Title": "", "Combined Score %": "0%", "Measurement Match %": "0%",
        "Vendor Measurements": "", "Amazon Measurements": "", "Match Alert": "",
        "Match Reason": "", "Match Recommendation": "", "Confidence Level": "LOW",
        "Action Required": "RE-EVALUATE", "Verification": "❌ FAILED", "Fail Reasons": "",
        "AI Insight": "", "Source": "scraper", "Cache Hit": "No",
    }
    
    # STEP 1: Determine which ASIN to use (UPC優先)
    asin_to_use = asin_from_sheet
    if upc and len(upc) >= 8:
        upc_result = find_single_item_asn(upc, session)
        if upc_result and upc_result.get("asin"):
            asin_to_use = upc_result["asin"]
            base["ASIN Source"] = f"UPC lookup ({upc_result['confidence']})"
            base["Pack Note"] = f"Pack size: {upc_result.get('pack_size', 1)}"
            if upc_result.get('pack_size', 1) != 1:
                base["Fail Reasons"] = f"⚠️ No single item found — using {upc_result.get('pack_size', 0)}-pack"
    
    base["ASIN Used"] = asin_to_use
    
    if not asin_to_use or len(asin_to_use) < 5:
        base["Verification"] = "⏭️ SKIPPED"
        return base
    
    # STEP 2: Fetch data
    cached = None if force_refresh else cache_get(asin_to_use)
    if cached:
        price_data, desc_data = cached["price_data"], cached["desc_data"]
        base["Cache Hit"] = "Yes"
    else:
        soup_product, status = fetch_asin_page(asin_to_use, session, config.get("max_retries", 3))
        if soup_product is None:
            base["Verification"] = "❌ FETCH FAILED"
            cache_mark_failed(asin_to_use, status)
            return base
        soup_offers = fetch_offer_listing(asin_to_use, session)
        seller_type = config.get("seller_type", "FBA")
        price_data = extract_all_prices(soup_product, soup_offers, your_bb, seller_type)
        desc_data = extract_product_text(soup_product)
        cache_set(asin_to_use, price_data, desc_data, {})
    
    # STEP 3: Enhanced description matching (unit-aware)
    amz_text = desc_data.get("full_text", "")
    match_result = enhanced_description_match(our_text, amz_text, sem_model)
    
    # STEP 4: Price logic
    sel_price = price_data.get("selected_price")
    max_price = price_data.get("max_price")
    bb_live = price_data.get("bb_price")
    price_range = price_data.get("price_range_str", "N/A")
    has_range = "–" in price_range
    
    # STEP 5: P&L
    pl = recalculate_pl(sel_price, max_price, net_price, full_cost, config.get("referral_rate", 0.15))
    
    # STEP 6: Confidence
    conf = compute_confidence(match_result['final_score'], "ok", sel_price is not None, has_range, True)
    
    # STEP 7: Verdict
    match_thresh = config.get("match_threshold", 0.35)
    hard_fails = []
    if match_result['final_score'] < match_thresh:
        hard_fails.append(f"Match too low ({match_result['final_score']*100:.0f}%) - {match_result['reason']}")
    if match_result['alert_level'] == "CRITICAL":
        hard_fails.append(match_result['reason'])
    if pl["status"] == "LOSS":
        hard_fails.append(f"Loss at current price")
    
    # STEP 8: Gemini AI Insight
    if st.session_state.get("gemini_configured", False) and match_result['final_score'] < 0.6:
        ai = ai_analyze_mismatch(our_text[:400], amz_text[:400], match_result['final_score'])
        base["AI Insight"] = ai.get("insight", "")
    
    # Final verdict
    if hard_fails:
        base["Verification"] = "❌ FAILED"
        base["Fail Reasons"] = " | ".join(hard_fails[:3])
    else:
        base["Verification"] = "✅ Verified"
    
    # Populate output
    base.update({
        "Price Range": price_range,
        "Selected Price": f"${sel_price:.2f}" if sel_price else "N/A",
        "FBA Prices": ", ".join(f"${p:.2f}" for p in sorted(set(price_data.get("fba_prices", [])))),
        "FBM Prices": ", ".join(f"${p:.2f}" for p in sorted(set(price_data.get("fbm_prices", [])))),
        "BB Price (Live)": f"${bb_live:.2f}" if bb_live else "N/A",
        "New Breakeven": f"${pl['new_breakeven']:.2f}" if pl['new_breakeven'] else "—",
        "Conservative Profit": f"${pl['conservative_profit']:.2f}" if pl['conservative_profit'] else "—",
        "Conservative Margin": f"{pl['conservative_margin']:.1f}%" if pl['conservative_margin'] else "—",
        "P&L Status": pl["status"],
        "Recommendation": pl["recommendation"],
        "Amazon Title": desc_data.get("title", "")[:100],
        "Combined Score %": f"{match_result['final_score']*100:.1f}%",
        "Measurement Match %": f"{match_result['measurement_score']*100:.1f}%",
        "Vendor Measurements": match_result['vendor_measurements'],
        "Amazon Measurements": match_result['amazon_measurements'],
        "Match Alert": match_result['alert_level'],
        "Match Reason": match_result['reason'],
        "Match Recommendation": match_result['recommendation'],
        "Confidence Level": conf["level"],
        "Action Required": conf["action"],
    })
    
    return base

# =============================================================================
# EXCEL EXPORT
# =============================================================================
def build_excel(df: pd.DataFrame, match_thresh: float) -> bytes:
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    
    NAVY, GREEN, RED, AMBER = [PatternFill("solid", fgColor=c) for c in ["1e2d4e", "C6EFCE", "FFC7CE", "FFEB9C"]]
    BOLD, WHITE = Font(bold=True), Font(bold=True, color="FFFFFF")
    WRAP, CTR = Alignment(wrap_text=True), Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment = NAVY, WHITE, CTR
    
    hdr = [c.value for c in ws[1]]
    def ci(n): return hdr.index(n) + 1 if n in hdr else None
    
    for row in ws.iter_rows(min_row=2):
        v_val = str(row[ci("Verification")-1].value or "") if ci("Verification") else ""
        if ci("Verification"):
            c = row[ci("Verification")-1]
            if "Verified" in v_val: c.fill, c.font = GREEN, BOLD
            elif "FAILED" in v_val: c.fill, c.font = RED, BOLD
    
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(mx + 3, 50)
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

# =============================================================================
# STREAMLIT UI
# =============================================================================
with st.sidebar:
    st.markdown("## ⚙️ Configuration")
    
    # Gemini API Section
    st.markdown("**🤖 Gemini AI (Free)**")
    gemini_key = st.text_input("Gemini API Key", type="password",
                               placeholder="Get free key from makersuite.google.com/app/apikey")
    if gemini_key and GEMINI_AVAILABLE:
        genai.configure(api_key=gemini_key)
        st.session_state["gemini_configured"] = True
        st.success("✅ Gemini AI Active")
    else:
        st.session_state["gemini_configured"] = False
        st.caption("Add free API key for AI mismatch analysis")
    
    st.markdown("---")
    
    # Column overrides
    st.markdown("**🗂️ Column Overrides**")
    ov_upc = st.text_input("UPC Column", placeholder="e.g. UPC, EAN, Barcode")
    ov_asin = st.text_input("ASIN Column", placeholder="e.g. Output ASIN")
    ov_title = st.text_input("Title Column")
    ov_desc = st.text_input("Description Column")
    ov_bb = st.text_input("BB Price Column")
    ov_net = st.text_input("Net Price Column")
    ov_full = st.text_input("Fulfillment Cost Col")
    
    st.markdown("---")
    MATCH_THRESH = st.slider("Min Match %", 10, 80, 35, 5) / 100
    N_WORKERS = st.slider("Parallel Workers", 1, 3, 2, 1)
    FORCE_REFRESH = st.checkbox("Force Refresh", value=False)

# =============================================================================
# MAIN APP
# =============================================================================
sem_model = load_semantic_model()

uploaded = st.file_uploader("Upload Excel", type=["xlsx", "xls"])

if uploaded:
    df = pd.read_excel(uploaded, dtype=str)
    df.columns = df.columns.str.strip()
    all_cols = list(df.columns)
    
    def res(ov, key): return ov if (ov and ov in all_cols) else detect_col(key, all_cols)
    
    UPC_COL = res(ov_upc, "upc")
    ASIN_COL = res(ov_asin, "asin")
    TITLE_COL = res(ov_title, "title")
    DESC_COL = res(ov_desc, "desc")
    BB_COL = res(ov_bb, "bb_price")
    NET_COL = res(ov_net, "net_price")
    FULL_COL = res(ov_full, "fulfillment")
    
    st.markdown('<p class="sec">🗂️ Column Detection</p>', unsafe_allow_html=True)
    cols_ui = st.columns(4)
    cols_ui[0].metric("UPC", "✅ " + UPC_COL[:15] if UPC_COL else "⚠️ Not found")
    cols_ui[1].metric("ASIN", "✅ " + ASIN_COL[:15] if ASIN_COL else "⚠️ Not found")
    cols_ui[2].metric("Title", "✅ " + TITLE_COL[:15] if TITLE_COL else "⚠️ Not found")
    cols_ui[3].metric("Description", "✅ " + DESC_COL[:15] if DESC_COL else "⚠️ Not found")
    
    if st.button(f"🚀 START VERIFICATION — {len(df)} Rows"):
        session = build_session(None)
        results = []
        progress = st.progress(0)
        
        for idx, row in df.iterrows():
            row_data = {
                "upc": str(row.get(UPC_COL, "")) if UPC_COL else "",
                "asin": str(row.get(ASIN_COL, "")) if ASIN_COL else "",
                "title": str(row.get(TITLE_COL, "")) if TITLE_COL else "",
                "desc": str(row.get(DESC_COL, "")) if DESC_COL else "",
                "brand": "",
                "your_bb": parse_price(row.get(BB_COL)) if BB_COL else None,
                "net_price": parse_price(row.get(NET_COL)) if NET_COL else None,
                "full_cost": parse_price(row.get(FULL_COL)) if FULL_COL else None,
            }
            result = verify_single_asin(row_data, session, sem_model,
                                        {"match_threshold": MATCH_THRESH, "referral_rate": 0.15},
                                        FORCE_REFRESH)
            results.append(result)
            progress.progress((idx + 1) / len(df))
        
        results_df = pd.DataFrame(results)
        for col in results_df.columns:
            if col not in df.columns:
                df[col] = results_df[col].values
        
        st.success(f"✅ Complete! {len(df)} ASINs verified")
        
        # Summary
        verified = df[df["Verification"].str.contains("Verified", na=False)].shape[0]
        failed = df[df["Verification"].str.contains("FAILED", na=False)].shape[0]
        st.metric("✅ Verified", verified)
        st.metric("❌ Failed", failed)
        
        excel_bytes = build_excel(df, MATCH_THRESH)
        st.download_button("📥 Download Excel Report", data=excel_bytes,
                          file_name="VirVentures_ASIN_Verification_v7.xlsx",
                          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
else:
    st.info("📂 Please upload an Excel file to begin")

st.markdown("""
<div style="text-align:center;padding:20px;font-size:12px;color:#aaa;">
    VirVentures ASIN Verifier v7.0 | Gemini AI · UPC Search · Pack Size Detection · Unit-Aware Matching
</div>
""", unsafe_allow_html=True)
